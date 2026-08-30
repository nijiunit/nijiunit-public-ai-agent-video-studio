from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from .assets import AssetRecord
from .schema import Shot, Storyboard

NAVY = "1F4E78"
BLUE = "D9EAF7"
PALE_BLUE = "EAF3F8"
YELLOW = "FFF2CC"
GREEN = "E2F0D9"
RED = "F4CCCC"
WHITE = "FFFFFF"
GRAY = "E7E6E6"
THIN = Side(style="thin", color="A6A6A6")
DEFAULT_CORRECTION_INSTRUCTION = "ここへ具体的な訂正指示を記入してください。"


def _review_image_size(aspect_ratio: str, long_edge: int) -> tuple[int, int]:
    if aspect_ratio == "9:16":
        return round(long_edge * 9 / 16), long_edge
    return long_edge, round(long_edge * 9 / 16)


def _timecode(seconds: float) -> str:
    minutes = int(seconds // 60)
    remainder = seconds - minutes * 60
    if abs(remainder - round(remainder)) < 0.0005:
        return f"{minutes:02d}:{round(remainder):02d}"
    return f"{minutes:02d}:{remainder:06.3f}"


def _safe_sheet_title(shot: Shot) -> str:
    clean = "".join(char for char in shot.title if char not in r"[]:*?/\\")
    return f"S{shot.shot_number:03d}_{clean}"[:31]


def _image_for_shot(
    run_dir: Path, shot: Shot, assets: list[AssetRecord]
) -> Path | None:
    generated = run_dir / "images" / f"shot_{shot.shot_number:03d}.png"
    if generated.exists():
        return generated
    requested = set(shot.reference_assets)
    for asset in assets:
        if asset.original_name in requested and asset.prepared_path:
            path = Path(asset.prepared_path)
            if path.exists():
                return path
    return None


def _set_common_dimensions(ws) -> None:
    for column in "ABCDEFGHIJKLMNOPQR":
        ws.column_dimensions[column].width = 12
    ws.column_dimensions["A"].width = 3
    for row in range(1, 55):
        ws.row_dimensions[row].height = 24


def _label(ws, cell: str, value: str) -> None:
    ws[cell] = value
    ws[cell].font = Font(bold=True, color=WHITE)
    ws[cell].fill = PatternFill("solid", fgColor=NAVY)
    ws[cell].alignment = Alignment(vertical="center")
    ws[cell].border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _value(ws, cell: str, value: str, fill: str = WHITE) -> None:
    ws[cell] = value
    ws[cell].fill = PatternFill("solid", fgColor=fill)
    ws[cell].alignment = Alignment(wrap_text=True, vertical="top")
    ws[cell].border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def create_workbook(
    storyboard: Storyboard,
    run_dir: Path,
    assets: list[AssetRecord],
    destination: Path,
) -> Path:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "00_全体"
    summary.sheet_view.showGridLines = False
    summary.freeze_panes = "A7"
    summary.column_dimensions["A"].width = 12
    summary.column_dimensions["B"].width = 22
    summary.column_dimensions["C"].width = 14
    summary.column_dimensions["D"].width = 65
    summary.column_dimensions["E"].width = 24

    summary.merge_cells("A1:E1")
    summary["A1"] = storyboard.title
    summary["A1"].font = Font(size=20, bold=True, color=WHITE)
    summary["A1"].fill = PatternFill("solid", fgColor=NAVY)
    summary["A1"].alignment = Alignment(horizontal="center")
    summary["A2"] = "ログライン"
    summary["B2"] = storyboard.logline
    summary.merge_cells("B2:E2")
    summary["A3"] = "総尺"
    summary["B3"] = f"{storyboard.total_duration_seconds}秒"
    summary["C3"] = "画角"
    summary["D3"] = storyboard.aspect_ratio
    summary["A4"] = "画作り"
    summary["B4"] = storyboard.visual_style
    summary.merge_cells("B4:E4")
    summary["A5"] = "使い方"
    summary["B5"] = (
        "各シートの黄色い訂正指示欄へ記入。小規模修正は「小規模」、"
        "全体の物語を変える場合は「大規模」を選択。"
    )
    summary.merge_cells("B5:E5")

    headers = ["シート", "時間", "状態", "内容", "修正規模"]
    for column, header in enumerate(headers, start=1):
        cell = summary.cell(row=7, column=column, value=header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center")

    for row, shot in enumerate(storyboard.shots, start=8):
        title = _safe_sheet_title(shot)
        summary.cell(row=row, column=1, value=f'=HYPERLINK("#\'{title}\'!A1","{title}")')
        summary.cell(
            row=row,
            column=2,
            value=f"{_timecode(shot.start_seconds)}–{_timecode(shot.end_seconds)}",
        )
        summary.cell(row=row, column=3, value="未確認")
        summary.cell(row=row, column=4, value=shot.scene_description)
        summary.cell(row=row, column=5, value="なし")
        for column in range(1, 6):
            summary.cell(row=row, column=column).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    for shot in storyboard.shots:
        ws = workbook.create_sheet(_safe_sheet_title(shot))
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A10"
        _set_common_dimensions(ws)

        ws.merge_cells("A1:R2")
        ws["A1"] = (
            f"S{shot.shot_number:03d}  {shot.title}    "
            f"{_timecode(shot.start_seconds)}–{_timecode(shot.end_seconds)}"
        )
        ws["A1"].font = Font(size=18, bold=True, color=WHITE)
        ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        _label(ws, "B4", "レビュー状態")
        ws.merge_cells("C4:F4")
        _value(ws, "C4", "未確認", GREEN)
        status_validation = DataValidation(
            type="list",
            formula1='"未確認,承認,修正必要"',
            allow_blank=False,
        )
        ws.add_data_validation(status_validation)
        status_validation.add(ws["C4"])

        _label(ws, "G4", "修正規模")
        ws.merge_cells("H4:J4")
        _value(ws, "H4", "なし", GREEN)
        scope_validation = DataValidation(
            type="list",
            formula1='"なし,小規模,大規模"',
            allow_blank=False,
        )
        ws.add_data_validation(scope_validation)
        scope_validation.add(ws["H4"])

        _label(ws, "B5", "訂正指示")
        ws.merge_cells("C5:R8")
        _value(
            ws,
            "C5",
            DEFAULT_CORRECTION_INSTRUCTION,
            YELLOW,
        )
        ws.conditional_formatting.add(
            "H4",
            CellIsRule(
                operator="equal",
                formula=['"大規模"'],
                fill=PatternFill("solid", fgColor=RED),
            ),
        )

        ws.merge_cells("B10:J27")
        ws["B10"] = "メイン画像"
        ws["B10"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B10"].fill = PatternFill("solid", fgColor=GRAY)
        image_path = _image_for_shot(run_dir, shot, assets)
        if image_path:
            image = ExcelImage(str(image_path))
            image.width, image.height = _review_image_size(
                storyboard.aspect_ratio,
                720,
            )
            ws.add_image(image, "B10")

        fields = [
            ("K10", "物語上の役割", "K11:R12", shot.story_purpose),
            ("K13", "画面説明", "K14:R16", shot.scene_description),
            ("K17", "登場", "K18:R18", "、".join(shot.characters)),
            ("K19", "動き", "K20:R21", shot.action),
            ("K22", "表情・感情", "K23:R23", shot.emotion),
            ("K24", "カメラ", "K25:R25", shot.camera),
            ("K26", "光・色", "K27:R27", shot.lighting),
        ]
        for label_cell, label, value_range, value in fields:
            _label(ws, label_cell, label)
            ws.merge_cells(value_range)
            _value(ws, value_range.split(":")[0], value, PALE_BLUE)

        _label(ws, "B29", "セリフ")
        ws.merge_cells("C29:H30")
        _value(ws, "C29", shot.dialogue or "なし")
        _label(ws, "I29", "ナレーション")
        ws.merge_cells("J29:R30")
        _value(ws, "J29", shot.narration or "なし")
        _label(ws, "B31", "音")
        ws.merge_cells("C31:H32")
        _value(ws, "C31", shot.sound)
        _label(ws, "I31", "連続性")
        ws.merge_cells("J31:R32")
        _value(ws, "J31", shot.continuity)

        ws.merge_cells("B34:R34")
        ws["B34"] = "動画生成後の9コマ（1秒3フレーム）"
        ws["B34"].font = Font(bold=True, color=WHITE)
        ws["B34"].fill = PatternFill("solid", fgColor=NAVY)
        ws["B34"].alignment = Alignment(horizontal="center")

        frame_width = 6
        for frame_index, description in enumerate(shot.frame_descriptions):
            grid_row = frame_index // 3
            grid_col = frame_index % 3
            start_col = 2 + grid_col * frame_width
            top_row = 36 + grid_row * 6
            end_col = start_col + frame_width - 1
            ws.merge_cells(
                start_row=top_row,
                start_column=start_col,
                end_row=top_row + 3,
                end_column=end_col,
            )
            cell = ws.cell(row=top_row, column=start_col)
            cell.value = (
                f"Frame {frame_index + 1}  "
                f"{shot.frame_offset_seconds(frame_index):.3f}s\n\n"
                f"{description}\n\n"
                "（動画生成後に画像を挿入）"
            )
            cell.fill = PatternFill("solid", fgColor=GRAY)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        ws.merge_cells("B54:R54")
        ws["B54"] = "参照素材: " + "、".join(shot.reference_assets)
        ws["B54"].alignment = Alignment(wrap_text=True)

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def extract_corrections(workbook_path: Path, destination: Path) -> Path:
    workbook = load_workbook(workbook_path, data_only=False)
    corrections: list[dict[str, str]] = []
    for ws in workbook.worksheets:
        if not ws.title.startswith("S"):
            continue
        instruction = str(ws["C5"].value or "").strip()
        if instruction == DEFAULT_CORRECTION_INSTRUCTION:
            instruction = ""
        corrections.append(
            {
                "sheet": ws.title,
                "review_status": str(ws["C4"].value or "未確認"),
                "revision_scope": str(ws["H4"].value or "なし"),
                "instruction": instruction,
            }
        )
    destination.write_text(
        json.dumps({"corrections": corrections}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return destination


def workbook_review_issues(
    storyboard: Storyboard,
    workbook_path: Path,
) -> list[str]:
    """Return reasons why an Excel storyboard is not ready for video generation."""
    if not workbook_path.exists():
        return [f"Excelコンテがありません: {workbook_path}"]

    workbook = load_workbook(workbook_path, data_only=False)
    issues: list[str] = []
    for shot in storyboard.shots:
        sheet_title = _safe_sheet_title(shot)
        if sheet_title not in workbook.sheetnames:
            issues.append(f"S{shot.shot_number:03d}: シートがありません")
            continue
        worksheet = workbook[sheet_title]
        status = str(worksheet["C4"].value or "未確認").strip()
        instruction = str(worksheet["C5"].value or "").strip()
        if status != "承認":
            issues.append(
                f"S{shot.shot_number:03d}: レビュー状態が「{status}」です"
            )
        if instruction and instruction != DEFAULT_CORRECTION_INSTRUCTION:
            issues.append(
                f"S{shot.shot_number:03d}: 未反映の訂正指示があります"
            )
    return issues


def approve_workbook(
    storyboard: Storyboard,
    workbook_path: Path,
) -> Path:
    """Mark every shot approved after the user explicitly approves the workbook."""
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excelコンテがありません: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=False)
    blocking: list[str] = []
    for shot in storyboard.shots:
        sheet_title = _safe_sheet_title(shot)
        if sheet_title not in workbook.sheetnames:
            blocking.append(f"S{shot.shot_number:03d}: シートがありません")
            continue
        worksheet = workbook[sheet_title]
        instruction = str(worksheet["C5"].value or "").strip()
        if instruction and instruction != DEFAULT_CORRECTION_INSTRUCTION:
            blocking.append(
                f"S{shot.shot_number:03d}: 訂正指示が残っています"
            )
        if str(worksheet["C4"].value or "").strip() == "修正必要":
            blocking.append(
                f"S{shot.shot_number:03d}: レビュー状態が「修正必要」です"
            )
    if blocking:
        raise RuntimeError(
            "Excelコンテを承認できません。先に修正を反映してください。\n"
            + "\n".join(blocking)
        )

    summary = workbook[workbook.sheetnames[0]]
    for row, shot in enumerate(storyboard.shots, start=8):
        sheet_title = _safe_sheet_title(shot)
        worksheet = workbook[sheet_title]
        worksheet["C4"] = "承認"
        worksheet["H4"] = "なし"
        summary.cell(row=row, column=3, value="承認")
        summary.cell(row=row, column=5, value="なし")
    workbook.save(workbook_path)
    return workbook_path


def create_video_workbook(
    storyboard: Storyboard,
    source_workbook: Path,
    run_dir: Path,
    destination: Path,
) -> Path:
    """Create a review workbook containing the nine real frames for every shot."""
    workbook = load_workbook(source_workbook)
    for shot in storyboard.shots:
        ws = workbook[_safe_sheet_title(shot)]
        for frame_index, description in enumerate(shot.frame_descriptions):
            grid_row = frame_index // 3
            grid_col = frame_index % 3
            start_col = 2 + grid_col * 6
            top_row = 36 + grid_row * 6
            frame_path = (
                run_dir
                / "frames"
                / f"shot_{shot.shot_number:03d}"
                / f"frame_{frame_index + 1:02d}.jpg"
            )
            if not frame_path.exists():
                raise FileNotFoundError(f"Missing extracted frame: {frame_path}")
            for row in range(top_row, top_row + 4):
                ws.row_dimensions[row].height = 48
            cell = ws.cell(row=top_row, column=start_col)
            cell.value = (
                f"Frame {frame_index + 1} | "
                f"{shot.frame_offset_seconds(frame_index):.3f}s"
            )
            cell.comment = Comment(description, "Gemini storyboard")
            image = ExcelImage(str(frame_path))
            image.width, image.height = _review_image_size(
                storyboard.aspect_ratio,
                400,
            )
            ws.add_image(image, cell.coordinate)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination
