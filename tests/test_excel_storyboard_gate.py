import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image

from video_storyboard import pipeline
from video_storyboard.schema import CharacterProfile, Shot, Storyboard
from video_storyboard.workbook import (
    approve_workbook,
    create_workbook,
    workbook_review_issues,
)


def make_storyboard() -> Storyboard:
    shots = [
        Shot(
            shot_number=number,
            title=f"shot {number}",
            story_purpose="test",
            scene_description="test scene",
            characters=[],
            action="small motion",
            emotion="calm",
            camera="locked",
            lighting="soft",
            sound="none",
            continuity="same scene",
            reference_assets=[],
            main_image_prompt="test",
            video_prompt="test",
            frame_descriptions=["frame"] * 9,
        )
        for number in range(1, 9)
    ]
    return Storyboard(
        title="test",
        logline="test",
        audience="test",
        visual_style="test",
        story_summary="test",
        character_bible=[CharacterProfile(name="test", description="test")],
        shots=shots,
    )


def prepare_workbook(tmp_path: Path) -> tuple[Storyboard, Path]:
    storyboard = make_storyboard()
    images = tmp_path / "images"
    images.mkdir()
    for shot in storyboard.shots:
        Image.new("RGB", (32, 18), "navy").save(
            images / f"shot_{shot.shot_number:03d}.png"
        )
    workbook_path = tmp_path / "storyboard_test.xlsx"
    create_workbook(storyboard, tmp_path, [], workbook_path)
    return storyboard, workbook_path


def test_excel_storyboard_requires_explicit_approval(tmp_path: Path) -> None:
    storyboard, workbook_path = prepare_workbook(tmp_path)

    issues = workbook_review_issues(storyboard, workbook_path)
    assert len(issues) == 8
    assert all("未確認" in issue for issue in issues)

    approve_workbook(storyboard, workbook_path)

    assert workbook_review_issues(storyboard, workbook_path) == []
    workbook = load_workbook(workbook_path, data_only=False)
    assert workbook[workbook.sheetnames[0]]["C8"].value == "承認"
    assert all(
        workbook[sheet]["C4"].value == "承認"
        for sheet in workbook.sheetnames[1:]
    )


def test_excel_storyboard_cannot_approve_pending_correction(
    tmp_path: Path,
) -> None:
    storyboard, workbook_path = prepare_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    workbook[workbook.sheetnames[1]]["C5"] = "主人公を左へ移動"
    workbook.save(workbook_path)

    with pytest.raises(RuntimeError, match="訂正指示が残っています"):
        approve_workbook(storyboard, workbook_path)


def test_video_generation_stops_before_excel_approval(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    storyboard = make_storyboard()
    (tmp_path / "storyboard.json").write_text(
        storyboard.model_dump_json(indent=2),
        encoding="utf-8",
    )
    called = False

    def fake_render_video_shots(**_kwargs):
        nonlocal called
        called = True
        return []

    monkeypatch.setattr(pipeline, "render_video_shots", fake_render_video_shots)

    with pytest.raises(RuntimeError, match="Excelコンテの確認と承認"):
        pipeline.render_videos_command(tmp_path, character_registry_dir=None)
    assert called is False


def test_workbook_build_stops_when_main_images_are_missing(
    tmp_path: Path,
) -> None:
    storyboard = make_storyboard()
    (tmp_path / "storyboard.json").write_text(
        storyboard.model_dump_json(indent=2),
        encoding="utf-8",
    )

    with pytest.raises(RuntimeError, match="全ショットのメイン画像"):
        pipeline.build_workbook_command(tmp_path)


def test_pipeline_builds_review_folder_html_and_new_revision(
    tmp_path: Path,
) -> None:
    run_dir = tmp_path / "v001"
    run_dir.mkdir()
    storyboard = make_storyboard()
    (run_dir / "storyboard.json").write_text(
        storyboard.model_dump_json(indent=2),
        encoding="utf-8",
    )
    (run_dir / "manifest.json").write_text(
        json.dumps({"assets": []}),
        encoding="utf-8",
    )
    image_dir = run_dir / "images"
    image_dir.mkdir()
    for shot in storyboard.shots:
        Image.new("RGB", (32, 18), "navy").save(
            image_dir / f"shot_{shot.shot_number:03d}.png"
        )

    first_result = pipeline.build_workbook_command(run_dir)
    first = run_dir / "review" / "storyboard_v001.xlsx"
    assert first.is_file()
    assert "確認用フォルダ" in first_result
    assert pipeline.review_html_path(first, "ja").is_file()
    assert pipeline.review_html_path(first, "en").is_file()
    corrections = Path(pipeline.extract_corrections_command(first))
    assert corrections.parent == run_dir
    assert corrections.name == "corrections_storyboard_v001.json"

    pipeline.build_workbook_command(run_dir)
    second = run_dir / "review" / "storyboard_v001_r002.xlsx"
    assert first.is_file()
    assert second.is_file()


def test_open_excel_lock_gets_one_clear_recovery_action(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    storyboard, workbook_path = prepare_workbook(tmp_path)
    (tmp_path / "storyboard.json").write_text(
        storyboard.model_dump_json(indent=2),
        encoding="utf-8",
    )

    monkeypatch.setattr(
        pipeline,
        "current_storyboard_workbook",
        lambda _run_dir: workbook_path,
    )

    def locked(*_args, **_kwargs):
        raise PermissionError("locked")

    monkeypatch.setattr(pipeline, "approve_workbook", locked)

    with pytest.raises(RuntimeError, match="保存してから閉じ"):
        pipeline.approve_workbook_command(tmp_path)
