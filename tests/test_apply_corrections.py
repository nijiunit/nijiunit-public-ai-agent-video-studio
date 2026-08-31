from __future__ import annotations

import json
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from PIL import Image

from video_storyboard import pipeline
from video_storyboard.schema import CharacterProfile, Shot, Storyboard
from video_storyboard.workbook import create_workbook


def _storyboard(description: str = "before") -> Storyboard:
    return Storyboard(
        title="test",
        logline="test",
        audience="test",
        visual_style="test",
        story_summary="test",
        character_bible=[CharacterProfile(name="none", description="none")],
        shots=[
            Shot(
                shot_number=1,
                title="shot",
                story_purpose="test",
                scene_description=description,
                characters=[],
                action="test",
                emotion="calm",
                camera="wide",
                lighting="soft",
                sound="none",
                continuity="same",
                reference_assets=[],
                main_image_prompt=description,
                video_prompt=description,
                frame_descriptions=[description] * 9,
            )
        ],
    )


def _two_shot_storyboard() -> Storyboard:
    first = _storyboard().shots[0]
    second = first.model_copy(
        update={
            "shot_number": 2,
            "title": "shot 2",
            "scene_description": "second",
            "main_image_prompt": "second",
            "video_prompt": "second",
            "frame_descriptions": ["second"] * 9,
            "continuity_start_mode": "previous_final_frame",
        },
        deep=True,
    )
    data = _storyboard().model_dump()
    data["shots"] = [first.model_dump(), second.model_dump()]
    return Storyboard.model_validate(data)


@pytest.mark.parametrize("correction_source", ["excel", "chat"])
def test_apply_corrections_revises_json_and_preserves_old_image(
    tmp_path: Path, monkeypatch, correction_source: str
) -> None:
    run_dir = tmp_path / "v001"
    image_dir = run_dir / "images"
    image_dir.mkdir(parents=True)
    storyboard = _storyboard()
    (run_dir / "storyboard.json").write_text(
        storyboard.model_dump_json(indent=2), encoding="utf-8"
    )
    (run_dir / "manifest.json").write_text(
        json.dumps({"assets": []}), encoding="utf-8"
    )
    Image.new("RGB", (32, 18), "navy").save(image_dir / "shot_001.png")
    workbook_path = run_dir / "review" / "storyboard_v001.xlsx"
    create_workbook(storyboard, run_dir, [], workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook[workbook.sheetnames[1]]
    sheet["C4"] = "修正必要"
    sheet["H4"] = "小規模"
    sheet["C5"] = "背景を明るくする"
    workbook.save(workbook_path)
    corrections_file = None
    if correction_source == "chat":
        corrections_file = tmp_path / "chat_corrections.json"
        corrections_file.write_text(
            json.dumps(
                {
                    "corrections": [
                        {
                            "sheet": "S001 shot",
                            "review_status": "修正必要",
                            "revision_scope": "小規模",
                            "instruction": "背景を明るくする",
                        }
                    ]
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

    guidance = SimpleNamespace(
        profile=SimpleNamespace(media=SimpleNamespace(aspect_ratio="16:9"))
    )
    monkeypatch.setattr(pipeline, "load_run_guidance", lambda _path: guidance)

    class FakeService:
        def __init__(self, **_kwargs):
            pass

        def revise_storyboard(self, *_args, **_kwargs):
            return _storyboard("after")

    monkeypatch.setattr(pipeline, "GeminiService", FakeService)

    result = pipeline.apply_corrections_command(
        run_dir,
        corrections_file=corrections_file,
        character_registry_dir=None,
    )

    target_run = tmp_path / "v002"
    original = Storyboard.model_validate_json(
        (run_dir / "storyboard.json").read_text(encoding="utf-8")
    )
    revised = Storyboard.model_validate_json(
        (target_run / "storyboard.json").read_text(encoding="utf-8")
    )
    assert original.shots[0].scene_description == "before"
    assert revised.shots[0].scene_description == "after"
    assert not list(target_run.glob("storyboard_r*.json"))
    assert (image_dir / "shot_001.png").is_file()
    assert not (target_run / "images" / "shot_001.png").exists()
    assert (
        target_run / "rejected" / "from_v001" / "images" / "shot_001.png"
    ).is_file()
    revision_record = json.loads(
        (target_run / "revision_origin.json").read_text(encoding="utf-8")
    )
    assert revision_record["corrections_source"] == (
        "chat_corrections.json"
        if correction_source == "chat"
        else "storyboard_v001.xlsx"
    )
    assert not (target_run / "review").exists()
    assert "再生成が必要な画像" in result
    assert str(target_run) in result


def test_correction_rejects_a_sheet_not_in_the_storyboard() -> None:
    data = {
        "corrections": [
            {
                "sheet": "S999 unknown",
                "review_status": "修正必要",
                "revision_scope": "小規模",
                "instruction": "背景を明るくする",
            }
        ]
    }

    with pytest.raises(RuntimeError, match="存在しません"):
        pipeline._validate_correction_data(data, _storyboard())


def test_small_correction_rejects_changes_outside_the_requested_shot() -> None:
    original = _two_shot_storyboard()
    revised = original.model_copy(deep=True)
    revised.shots[1].scene_description = "unexpected change"
    requested = [
        {
            "sheet": "S001 shot",
            "review_status": "修正必要",
            "revision_scope": "小規模",
            "instruction": "S001の背景を明るくする",
        }
    ]

    with pytest.raises(RuntimeError, match="対象外ショット.*S002"):
        pipeline._validate_revised_storyboard_scope(original, revised, requested)


def test_video_revision_expands_previous_frame_dependencies() -> None:
    storyboard = _two_shot_storyboard()

    assert pipeline._expand_video_continuity_dependents(storyboard, (1,)) == (1, 2)
