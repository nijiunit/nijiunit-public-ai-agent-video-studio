from pathlib import Path

from openpyxl import load_workbook

from video_storyboard.schema import Storyboard
from video_storyboard.workbook import workbook_review_issues

ROOT = Path(__file__).resolve().parents[1]


def test_public_storyboard_is_ten_three_second_shots() -> None:
    storyboard = Storyboard.model_validate_json(
        (ROOT / "examples" / "space-friends" / "storyboard.json").read_text(
            encoding="utf-8"
        )
    )
    assert len(storyboard.shots) == 10
    assert storyboard.total_duration_seconds == 30
    storyboard_image_shots = {
        shot.shot_number
        for shot in storyboard.shots
        if shot.continuity_start_mode == "storyboard_image"
    }
    assert storyboard_image_shots == {1, 4, 6, 8, 9}


def test_public_sample_includes_approved_excel_storyboard() -> None:
    storyboard = Storyboard.model_validate_json(
        (ROOT / "examples" / "space-friends" / "storyboard.json").read_text(
            encoding="utf-8"
        )
    )
    workbook_path = (
        ROOT / "examples" / "space-friends" / "storyboard_approved.xlsx"
    )

    assert workbook_path.is_file()
    assert workbook_review_issues(storyboard, workbook_path) == []
    workbook = load_workbook(workbook_path, data_only=False)
    assert len(workbook.sheetnames) == 11
    assert sum(len(workbook[sheet]._images) for sheet in workbook.sheetnames[1:]) == 10


def test_public_sample_has_offline_bilingual_storyboard_review() -> None:
    sample = ROOT / "examples" / "space-friends"
    japanese = sample / "storyboard_approved_review.ja.html"
    english = sample / "storyboard_approved_review.en.html"

    assert japanese.is_file()
    assert english.is_file()
    assert japanese.read_text(encoding="utf-8").count('class="shot"') == 10
    assert english.read_text(encoding="utf-8").count('class="shot"') == 10
    assert "http://" not in japanese.read_text(encoding="utf-8")
    assert "https://" not in english.read_text(encoding="utf-8")
    assets = sample / "storyboard_approved_review_assets"
    assert len(list(assets.glob("shot_*_main.jpg"))) == 10
